from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from reporting.reporting import (
    export_applicant_information,
    export_assess_feature_stats,
    export_end_of_application_survey_data,
)


@patch("reporting.reporting.get_funds_with_rounds")
@patch("reporting.reporting.get_assessments_by_round")
def test_export_applicant_information(
    mock_get_assessments_by_round: MagicMock,
    mock_get_funds_with_rounds: MagicMock,
) -> None:
    mock_fund = MagicMock()
    mock_fund.name_json = {"en": "Fund A"}
    mock_fund.short_name = "FA"

    mock_round = MagicMock()
    mock_round.title_json = {"en": "Round 1"}
    mock_round.short_name = "R1"
    mock_fund.rounds = [mock_round]

    mock_get_funds_with_rounds.return_value = [mock_fund]

    mock_assessment = MagicMock()
    mock_assessment.application_id = 123
    mock_assessment.jsonb_blob = {"date_submitted": "2024-03-11T10:00:00"}

    mock_get_assessments_by_round.return_value = [mock_assessment]

    workbook = Workbook()

    export_applicant_information(workbook)

    sheet = workbook["Applicant information"]

    assert sheet.cell(row=2, column=1).value == "Fund A"
    assert sheet.cell(row=2, column=2).value == "Fund A Round 1"
    assert sheet.cell(row=2, column=3).value == "FA-R1"
    assert sheet.cell(row=2, column=4).value == "123"
    assert sheet.cell(row=2, column=5).value == "2024-03-11 10:00:00"
    assert sheet.cell(row=2, column=6).value == "Apply"


@patch("reporting.reporting.get_funds_with_rounds")
@patch("reporting.reporting.get_applications_for_round_by_status")
def test_export_end_of_application_survey_data(
    mock_get_applications_for_round_by_status: MagicMock,
    mock_get_funds_with_rounds: MagicMock,
) -> None:
    mock_fund = MagicMock()
    mock_fund.short_name = "FA"

    mock_round = MagicMock()
    mock_round.short_name = "R1"
    mock_fund.rounds = [mock_round]

    mock_get_funds_with_rounds.return_value = [mock_fund]

    mock_application = MagicMock()
    mock_application.id = 123

    mock_feedback = MagicMock()
    mock_feedback.date_submitted = "2024-03-11T10:00:00"
    mock_feedback.get_section_comment_rating = ("Section 1", "Great feedback", 5)

    mock_application.end_of_application_survey = [mock_feedback]

    mock_get_applications_for_round_by_status.return_value = [mock_application]

    workbook = Workbook()

    export_end_of_application_survey_data(workbook)

    sheet = workbook["End of application survey data"]

    assert sheet.cell(row=2, column=1).value == "FA-R1"
    assert sheet.cell(row=2, column=2).value == "123"
    assert sheet.cell(row=2, column=3).value == "Section 1"
    assert sheet.cell(row=2, column=4).value == "Great feedback"
    assert sheet.cell(row=2, column=5).value == 5
    assert sheet.cell(row=2, column=6).value == "2024-03-11 10:00:00"


@patch("reporting.reporting.get_funds_with_rounds")
@patch("reporting.reporting.get_applications_for_round_by_status")
@patch("reporting.reporting.get_assessments_by_round")
@patch("reporting.reporting.get_applications_stats")
@patch("reporting.reporting.get_assessments_stats")
@patch("reporting.reporting.get_assessment_averages")
def test_export_assess_feature_stats(
    mock_get_assessment_averages: MagicMock,
    mock_get_assessments_stats: MagicMock,
    mock_get_applications_stats: MagicMock,
    mock_get_assessments_by_round: MagicMock,
    mock_get_applications_for_round_by_status: MagicMock,
    mock_get_funds_with_rounds: MagicMock,
) -> None:
    mock_fund = MagicMock()
    mock_fund.name_json = {"en": "Fund A"}
    mock_fund.short_name = "FA"

    mock_round = MagicMock()
    mock_round.title_json = {"en": "Round 1"}
    mock_round.short_name = "R1"
    mock_fund.rounds = [mock_round]

    mock_get_funds_with_rounds.return_value = [mock_fund]

    mock_application = MagicMock()
    mock_get_applications_for_round_by_status.return_value = [mock_application]
    mock_get_applications_stats.return_value = {
        "not_started": 1,
        "in_progress": 2,
        "completed": 3,
        "submitted": 4,
    }

    mock_assessment = MagicMock()
    mock_get_assessments_by_round.return_value = [mock_assessment]
    mock_get_assessments_stats.return_value = {
        "completed": 5,
        "not_started": 6,
        "in_progress": 7,
        "withdrawn": 8,
        "comments": 9,
        "tags": 10,
        "flags": 11,
        "change_requests": 12,
    }

    mock_get_assessment_averages.return_value = {
        "application_success_rate": 50,
        "comments_per_assessment": 1.5,
        "tags_per_assessment": 2.0,
        "flags_per_assessment": 0.5,
        "change_requests_per_assessment": 0.8,
    }

    workbook = Workbook()

    export_assess_feature_stats(workbook)

    sheet = workbook["Assess feature stats"]

    assert sheet.cell(row=2, column=1).value == "Fund A Round 1"
    assert sheet.cell(row=2, column=2).value == "FA-R1"
    assert sheet.cell(row=2, column=3).value == 1  # Number of applications not started
    assert sheet.cell(row=2, column=4).value == 2  # Number of applications in progress
    assert sheet.cell(row=2, column=5).value == 3  # Number of applications completed but not submitted
    assert sheet.cell(row=2, column=6).value == 4  # Number of applications submitted
    assert sheet.cell(row=2, column=7).value == 1  # Number of assessments received
    assert sheet.cell(row=2, column=8).value == 5  # Number of assessments completed
    assert sheet.cell(row=2, column=9).value == 6  # Number of assessments not started
    assert sheet.cell(row=2, column=10).value == 7  # Number of assessments in progress
    assert sheet.cell(row=2, column=11).value == 8  # Number of assessments withdrawn
    assert sheet.cell(row=2, column=12).value == "50%"  # Application success rate
    assert sheet.cell(row=2, column=13).value == 9  # Total assessment comments
    assert sheet.cell(row=2, column=14).value == 1.5  # Comments per assessment
    assert sheet.cell(row=2, column=15).value == 10  # Total tags assigned
    assert sheet.cell(row=2, column=16).value == 2.0  # Tags per assessment
    assert sheet.cell(row=2, column=17).value == 11  # Total assessment flags
    assert sheet.cell(row=2, column=18).value == 0.5  # Flags per assessment
    assert sheet.cell(row=2, column=19).value == 12  # Total change requests
    assert sheet.cell(row=2, column=20).value == 0.8  # Change requests per assessment
