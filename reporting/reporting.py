from datetime import datetime

from openpyxl import Workbook

from data.crud.applications import get_applications_for_round_by_status
from data.crud.assessment_records import get_assessments_by_round
from data.crud.fund_round_queries import get_funds_with_rounds
from pre_award.application_store.db.models.application.enums import Status as ApplicationStatus
from reporting.helpers import get_applications_stats, get_assessment_averages, get_assessments_stats


def export_applicant_information(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(title="Applicant information")

    sheet.append(["Fund", "Fund Round", "FundID", "ID", "Submission date", "Product"])

    funds = get_funds_with_rounds()
    for fund in funds:
        for round in fund.rounds:
            assessments = get_assessments_by_round(round.id)
            for assessment in assessments:
                date_submitted = datetime.fromisoformat(str(assessment.jsonb_blob["date_submitted"]))

                row = [
                    fund.name_json["en"],
                    fund.name_json["en"] + " " + round.title_json["en"],
                    fund.short_name + "-" + round.short_name,
                    str(assessment.application_id),
                    date_submitted.strftime("%Y-%m-%d %H:%M:%S"),
                    "Apply",
                ]

                sheet.append(row)


def export_end_of_application_survey_data(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(title="End of application survey data")

    sheet.append(["Fund", "application_id", "section", "comment", "rating", "date_submitted"])

    funds = get_funds_with_rounds()
    for fund in funds:
        for round in fund.rounds:
            applications = get_applications_for_round_by_status(round.id, [ApplicationStatus.SUBMITTED])
            for application in applications:
                feedback = application.end_of_application_survey
                for item in feedback:
                    date_submitted = datetime.fromisoformat(str(item.date_submitted))
                    section, comment, rating = item.get_section_comment_rating

                    row = [
                        fund.short_name + "-" + round.short_name,
                        str(application.id),
                        section,
                        comment,
                        rating,
                        date_submitted.strftime("%Y-%m-%d %H:%M:%S"),
                    ]

                    sheet.append(row)


def export_assess_feature_stats(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(title="Assess feature stats")

    sheet.append(
        [
            "Fund Round",
            "Fund-Round",
            "Applications not started",
            "Applications in progress",
            "Applications completed but not submitted",
            "Applications submitted",
            "Assessments received",
            "Assessments completed",
            "Assessments not started",
            "Assessments in progress",
            "Assessments withdrawn",
            "Application success rate %",
            "Total assessment comments",
            "Comments per assessment",
            "Total tags assigned",
            "Tags per assessment",
            "Total assessment flags",
            "Flags per assessment",
            "Total change requests",
            "Change requests per assessment",
        ]
    )

    funds = get_funds_with_rounds()
    for fund in funds:
        for round in fund.rounds:
            applications = get_applications_for_round_by_status(round.id)
            application_stats = get_applications_stats(applications)
            assessments = get_assessments_by_round(round.id)
            assessment_stats = get_assessments_stats(assessments)
            assessment_averages = get_assessment_averages(assessments, assessment_stats)

            row = {
                "Fund Round": fund.name_json["en"] + " " + round.title_json["en"],
                "Fund-Round": fund.short_name + "-" + round.short_name,
                "Applications not started": application_stats["not_started"],
                "Applications in progress": application_stats["in_progress"],
                "Applications completed but not submitted": application_stats["completed"],
                "Applications submitted": application_stats["submitted"],
                "Assessments received": len(assessments),
                "Assessments completed": assessment_stats["completed"],
                "Assessments not started": assessment_stats["not_started"],
                "Assessments in progress": assessment_stats["in_progress"],
                "Assessments withdrawn": assessment_stats["withdrawn"],
                "Application success rate": f"{assessment_averages['application_success_rate']}%",
                "Total assessment comments": assessment_stats["comments"],
                "Comments per assessment": assessment_averages["comments_per_assessment"],
                "Total tags assigned": assessment_stats["tags"],
                "Tags per assessment": assessment_averages["tags_per_assessment"],
                "Total assessment flags": assessment_stats["flags"],
                "Flags per assessment": assessment_averages["flags_per_assessment"],
                "Total change requests": assessment_stats["change_requests"],
                "Change requests per assessment": assessment_averages["change_requests_per_assessment"],
            }

            sheet.append(list(row.values()))
